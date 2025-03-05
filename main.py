from flask import Flask, render_template, request, redirect
import openpyxl
import os


app = Flask(__name__)

data_file = "server_testing.xlsx"
sheet_name = "Tháng 02.2025"


@app.route('/')
def index():
    return render_template('form.html')


@app.route('/submit', methods=['POST'])
def submit():
    wb = openpyxl.load_workbook(data_file)
    sheet = wb[sheet_name]

    # ✅ Find the next available column (starting from Column C)
    next_column = 3
    while sheet.cell(row=1, column=next_column).value is not None:
        next_column += 2  # ✅ Skips one column after every submission

    # ✅ Define form fields and their expected formats
    number_fields = {"actual_weight", "post_fat_removal_weight", "pieces_per_kg","ingredient", "actual_weight", "post_fat_removal_weight",
        "defective_material", "bong_1", "bong_so", "bong_cat", "bong_8_10", "bong_less_8", "bong_b",
        "duoi", "vun", "mo", "dat", "bong_1_2", "bong_so_2", "bong_la", "bong_la_b", "dat_2", "vun_2",
        "la_total", "la_1", "la_1_vang", "la_2", "la_2_vang", "la_3", "la_3_vang", "la_4", "la_4_vang",
        "la_5", "la_5_6", "la_5_vang", "la_6", "la_6_vang", "la_7", "la_7_vang", "la_8", "la_8_vang",
        "la_8_9", "la_9", "la_9_vang", "la_10", "la_10_vang", "la_11", "la_11_vang", "la_12", "la_12_vang",
        "la_sample", "la_yellow", "la_green", "la_yellow_no_pass", "la_yellow_bad_smell_pass",
        "la_yellow_bad_smell", "la_white", "la_puff", "la_no_pass", "lb_total", "lb_1", "lb_1_vang",
        "lb_60_80", "lb_80_120", "lb_120_160", "lb_2", "lb_2_vang", "lb_3", "lb_3_vang",
        "lb_200_280", "lb_4", "lb_4_vang", "lb_5", "lb_5_6", "lb_5_vang", "lb_6", "lb_6_vang",
        "lb_7", "lb_7_vang", "lb_8", "lb_8_vang", "lb_8_9", "lb_9", "lb_480_760", "lb_9_vang",
        "lb_10", "lb_10_vang", "lb_10_vang_1", "lb_760_1000", "lb_sample", "lb_piece", "lb_yellow",
        "lb_green", "lb_yellow_sample", "lb_yellow_no_pass_bad_smell", "lb_yellow_bad_smell_pass",
        "lb_white", "lb_white_bad_smell", "lb_yellow_no_pass", "lb_yellow_bad_smell", "lb_no_pass",
        "lc_total", "lc_1", "lc_1_vang", "lc_1_good", "lc_0", "lc_0_vang", "lc_2", "lc_2_vang",
        "lc_3", "lc_3_vang", "lc_4", "lc_4_vang", "lc_5", "lc_5_vang", "lc_6", "lc_7", "lc_8",
        "lc_9", "lc_10", "lc_10_vang", "lc_yellow", "lc_sample", "lc_good", "lc_good_yellow",
        "lc_white", "lc_bad_smell_white", "lc_yellow_hole", "lc_yellow_no_pass", "lc_green",
        "lc_yellow_bad_smell_pass", "lc_black_no_pass", "lc_black_bad_smell_no_pass", "lc_no_pass"}
    percentage_fields = {"fat_removal_date"}  # Example field
    currency_fields = {"cost", "total_price"}
    date_fields = {"entry_date", "production_date", "fat_removal_date"}

    form_fields = [
        "entry_date", "supplier", "production_date", "pieces_per_kg", "fat_removal_date",
        "raw_material_code", "remarks", "ingredient", "actual_weight", "post_fat_removal_weight",
        "defective_material", "bong_1", "bong_so", "bong_cat", "bong_8_10", "bong_less_8", "bong_b",
        "duoi", "vun", "mo", "dat", "bong_1_2", "bong_so_2", "bong_la", "bong_la_b", "dat_2", "vun_2",
        "la_total", "la_1", "la_1_vang", "la_2", "la_2_vang", "la_3", "la_3_vang", "la_4", "la_4_vang",
        "la_5", "la_5_6", "la_5_vang", "la_6", "la_6_vang", "la_7", "la_7_vang", "la_8", "la_8_vang",
        "la_8_9", "la_9", "la_9_vang", "la_10", "la_10_vang", "la_11", "la_11_vang", "la_12", "la_12_vang",
        "la_sample", "la_yellow", "la_green", "la_yellow_no_pass", "la_yellow_bad_smell_pass",
        "la_yellow_bad_smell", "la_white", "la_puff", "la_no_pass", "lb_total", "lb_1", "lb_1_vang",
        "lb_60_80", "lb_80_120", "lb_120_160", "lb_2", "lb_2_vang", "lb_3", "lb_3_vang",
        "lb_200_280", "lb_4", "lb_4_vang", "lb_5", "lb_5_6", "lb_5_vang", "lb_6", "lb_6_vang",
        "lb_7", "lb_7_vang", "lb_8", "lb_8_vang", "lb_8_9", "lb_9", "lb_480_760", "lb_9_vang",
        "lb_10", "lb_10_vang", "lb_10_vang_1", "lb_760_1000", "lb_sample", "lb_piece", "lb_yellow",
        "lb_green", "lb_yellow_sample", "lb_yellow_no_pass_bad_smell", "lb_yellow_bad_smell_pass",
        "lb_white", "lb_white_bad_smell", "lb_yellow_no_pass", "lb_yellow_bad_smell", "lb_no_pass",
        "lc_total", "lc_1", "lc_1_vang", "lc_1_good", "lc_0", "lc_0_vang", "lc_2", "lc_2_vang",
        "lc_3", "lc_3_vang", "lc_4", "lc_4_vang", "lc_5", "lc_5_vang", "lc_6", "lc_7", "lc_8",
        "lc_9", "lc_10", "lc_10_vang", "lc_yellow", "lc_sample", "lc_good", "lc_good_yellow",
        "lc_white", "lc_bad_smell_white", "lc_yellow_hole", "lc_yellow_no_pass", "lc_green",
        "lc_yellow_bad_smell_pass", "lc_black_no_pass", "lc_black_bad_smell_no_pass", "lc_no_pass"
    ]

    # ✅ Retrieve form data safely using `.get()`
    form_data = {field: request.form.get(field, "") for field in form_fields}

    # ✅ Define skipped rows
    skipped_rows = {8, 29, 30, 67, 110, 144}
    current_row = 1

    # ✅ Write data while skipping specified rows
    for field, value in form_data.items():
        while current_row in skipped_rows:
            current_row += 1

        cell = sheet.cell(row=current_row, column=next_column)

        # ✅ Convert data format before writing to Excel
        if value:
            try:
                if field in number_fields:
                    cell.value = float(value)
                    cell.number_format = '0.00'  # Decimal format
                elif field in percentage_fields:
                    cell.value = float(value) / 100  # Convert to fraction for percentage
                    cell.number_format = '0.00%'
                elif field in currency_fields:
                    cell.value = float(value)
                    cell.number_format = '"$"#,##0.00'  # Currency format
                elif field in date_fields:
                    cell.value = datetime.strptime(value, "%Y-%m-%d")  # Convert string to date
                    cell.number_format = 'YYYY-MM-DD'  # Standard Date Format
                else:
                    cell.value = value  # Default format
            except ValueError:
                cell.value = value  # If conversion fails, store as text
        current_row += 1

    wb.save(data_file)
    print(f"✅ Data saved to Column {next_column}, skipped rows: {skipped_rows}")
    return redirect('/')


if __name__ == '__main__':
    app.run(debug=True, port=5001)
